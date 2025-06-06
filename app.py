#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Tracking BRIX - Versão Web
Desenvolvido com Streamlit para acesso online
Escritório de contabilidade - Brasil
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Configuração da página
st.set_page_config(
    page_title="🚢 Sistema BRIX - Tracking de Trânsito",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #2c3e50 0%, #3498db 100%);
        padding: 2rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        border-left: 4px solid #3498db;
    }
    .verde-card {
        border-left-color: #27ae60 !important;
    }
    .vermelho-card {
        border-left-color: #e74c3c !important;
    }
    .stButton > button {
        width: 100%;
        border-radius: 8px;
        font-weight: bold;
        transition: all 0.3s ease;
    }
    .verde-btn {
        background-color: #27ae60;
        color: white;
    }
    .vermelho-btn {
        background-color: #e74c3c;
        color: white;
    }
    .azul-btn {
        background-color: #3498db;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

# Dados da empresa
DADOS_EMPRESA = {
    'nome': 'BRIX LOGÍSTICA',
    'endereco': 'Rua das Flores, 123 - Centro',
    'cidade': 'Curitiba - PR',
    'telefone': '(41) 3333-4444',
    'email': 'contato@brixlogistica.com.br',
    'cnpj': '12.345.678/0001-90'
}

# Colunas do sistema
COLUNAS = [
    'CLIENTE', 'CONTAINER', 'CARREGAMENTO', 'EMBARQUE NAVIO',
    'SAIDA NAVIO', 'PREVISAO CHEGADA PARANAGUA', 'CHEGADA PARANAGUA',
    'CANAL RFB', 'LIBERAÇAO PARANAGUA', 'CHEGADA CIUDAD DEL ESTE PY',
    'DESCARREGAMENTO'
]

# Inicializar dados na sessão
if 'df_tracking' not in st.session_state:
    st.session_state.df_tracking = pd.DataFrame(columns=COLUNAS)

# Função para criar dados de exemplo
def criar_dados_exemplo():
    dados_exemplo = [
        {
            'CLIENTE': 'EMPRESA ABC LTDA',
            'CONTAINER': 'TCLU1234567',
            'CARREGAMENTO': '15/05/2025',
            'EMBARQUE NAVIO': '18/05/2025',
            'SAIDA NAVIO': '20/05/2025',
            'PREVISAO CHEGADA PARANAGUA': '25/05/2025',
            'CHEGADA PARANAGUA': '24/05/2025',
            'CANAL RFB': 'VERDE',
            'LIBERAÇAO PARANAGUA': '24/05/2025',
            'CHEGADA CIUDAD DEL ESTE PY': '26/05/2025',
            'DESCARREGAMENTO': '28/05/2025'
        },
        {
            'CLIENTE': 'COMERCIAL XYZ S.A.',
            'CONTAINER': 'MSKU9876543',
            'CARREGAMENTO': '20/05/2025',
            'EMBARQUE NAVIO': '23/05/2025',
            'SAIDA NAVIO': '25/05/2025',
            'PREVISAO CHEGADA PARANAGUA': '30/05/2025',
            'CHEGADA PARANAGUA': '29/05/2025',
            'CANAL RFB': 'VERMELHO',
            'LIBERAÇAO PARANAGUA': '',
            'CHEGADA CIUDAD DEL ESTE PY': '',
            'DESCARREGAMENTO': ''
        },
        {
            'CLIENTE': 'IMPORTADORA DEF',
            'CONTAINER': 'HLBU5555555',
            'CARREGAMENTO': '10/05/2025',
            'EMBARQUE NAVIO': '12/05/2025',
            'SAIDA NAVIO': '14/05/2025',
            'PREVISAO CHEGADA PARANAGUA': '20/05/2025',
            'CHEGADA PARANAGUA': '19/05/2025',
            'CANAL RFB': 'VERDE',
            'LIBERAÇAO PARANAGUA': '19/05/2025',
            'CHEGADA CIUDAD DEL ESTE PY': '21/05/2025',
            'DESCARREGAMENTO': '22/05/2025'
        }
    ]
    return pd.DataFrame(dados_exemplo)

# Função para colorir linhas baseado no canal RFB
def colorir_linha(row):
    if row['CANAL RFB'] == 'VERDE':
        return ['background-color: #d5f4e6'] * len(row)
    elif row['CANAL RFB'] == 'VERMELHO':
        return ['background-color: #fadbd8'] * len(row)
    else:
        return [''] * len(row)

# Função para exportar Excel
def exportar_excel(df):
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Tracking BRIX', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Tracking BRIX']
        
        # Formatação do cabeçalho
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Colorir linhas baseado no canal RFB
        verde_fill = PatternFill(start_color="d5f4e6", end_color="d5f4e6", fill_type="solid")
        vermelho_fill = PatternFill(start_color="fadbd8", end_color="fadbd8", fill_type="solid")
        
        for row_num, row_data in enumerate(df.itertuples(), 2):
            canal_rfb = getattr(row_data, 'CANAL_RFB', '')
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if canal_rfb == 'VERDE':
                    cell.fill = verde_fill
                elif canal_rfb == 'VERMELHO':
                    cell.fill = vermelho_fill
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    buffer.seek(0)
    return buffer

# Interface principal
def main():
    # Cabeçalho
    st.markdown(f"""
    <div class="main-header">
        <h1>🚢 {DADOS_EMPRESA['nome']}</h1>
        <h3>Sistema de Tracking de Trânsito Online</h3>
        <p>📍 {DADOS_EMPRESA['endereco']} - {DADOS_EMPRESA['cidade']}</p>
        <p>📞 {DADOS_EMPRESA['telefone']} | 📧 {DADOS_EMPRESA['email']}</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.header("🔧 Controles")
        
        # Carregar dados de exemplo
        if st.button("📋 Carregar Dados de Exemplo", type="primary"):
            st.session_state.df_tracking = criar_dados_exemplo()
            st.success("✅ Dados de exemplo carregados!")
            st.rerun()
        
        # Upload de arquivo
        st.subheader("📂 Importar Excel")
        uploaded_file = st.file_uploader("Escolha um arquivo Excel", type=['xlsx', 'xls'])
        
        if uploaded_file is not None:
            try:
                df_uploaded = pd.read_excel(uploaded_file)
                colunas_faltando = set(COLUNAS) - set(df_uploaded.columns)
                
                if colunas_faltando:
                    st.error(f"❌ Colunas faltando: {', '.join(colunas_faltando)}")
                else:
                    if st.button("📥 Importar Dados"):
                        st.session_state.df_tracking = df_uploaded[COLUNAS].copy()
                        st.success("✅ Dados importados com sucesso!")
                        st.rerun()
            except Exception as e:
                st.error(f"❌ Erro ao ler arquivo: {str(e)}")
        
        # Download dos dados
        if not st.session_state.df_tracking.empty:
            st.subheader("💾 Exportar Dados")
            buffer = exportar_excel(st.session_state.df_tracking)
            
            st.download_button(
                label="📊 Baixar Excel Colorido",
                data=buffer,
                file_name=f"tracking_brix_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # Dashboard principal
    if st.session_state.df_tracking.empty:
        st.warning("⚠️ Nenhum dado encontrado. Use os controles da barra lateral para carregar dados.")
        return
    
    # Métricas principais
    col1, col2, col3, col4 = st.columns(4)
    
    total_registros = len(st.session_state.df_tracking)
    verde_count = len(st.session_state.df_tracking[st.session_state.df_tracking['CANAL RFB'] == 'VERDE'])
    vermelho_count = len(st.session_state.df_tracking[st.session_state.df_tracking['CANAL RFB'] == 'VERMELHO'])
    pendentes = len(st.session_state.df_tracking[st.session_state.df_tracking['CANAL RFB'].isin(['', None])])
    
    with col1:
        st.metric("📦 Total de Containers", total_registros)
    
    with col2:
        st.metric("🟢 Canal Verde", verde_count, delta=f"{(verde_count/total_registros*100):.1f}%" if total_registros > 0 else "0%")
    
    with col3:
        st.metric("🔴 Canal Vermelho", vermelho_count, delta=f"{(vermelho_count/total_registros*100):.1f}%" if total_registros > 0 else "0%")
    
    with col4:
        st.metric("⏳ Pendentes", pendentes)
    
    # Gráficos
    col1, col2 = st.columns(2)
    
    with col1:
        # Gráfico de pizza - Canal RFB
        canal_counts = st.session_state.df_tracking['CANAL RFB'].value_counts()
        if not canal_counts.empty:
            fig_pie = px.pie(
                values=canal_counts.values,
                names=canal_counts.index,
                title="📊 Distribuição por Canal RFB",
                color_discrete_map={'VERDE': '#27ae60', 'VERMELHO': '#e74c3c', '': '#95a5a6'}
            )
            fig_pie.update_layout(height=400)
            st.plotly_chart(fig_pie, use_container_width=True)
    
    with col2:
        # Gráfico de barras - Clientes
        cliente_counts = st.session_state.df_tracking['CLIENTE'].value_counts().head(10)
        if not cliente_counts.empty:
            fig_bar = px.bar(
                x=cliente_counts.values,
                y=cliente_counts.index,
                orientation='h',
                title="📈 Top 10 Clientes",
                color_discrete_sequence=['#3498db']
            )
            fig_bar.update_layout(height=400, yaxis={'categoryorder': 'total ascending'})
            st.plotly_chart(fig_bar, use_container_width=True)
    
    # Filtros
    st.subheader("🔍 Filtros")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        filtro_cliente = st.text_input("Cliente", placeholder="Digite o nome do cliente...")
    
    with col2:
        filtro_container = st.text_input("Container", placeholder="Digite o número do container...")
    
    with col3:
        filtro_canal = st.selectbox("Canal RFB", ['Todos', 'VERDE', 'VERMELHO'])
    
    with col4:
        st.write("")  # Espaçamento
        if st.button("🔄 Limpar Filtros"):
            st.rerun()
    
    # Aplicar filtros
    df_filtrado = st.session_state.df_tracking.copy()
    
    if filtro_cliente:
        df_filtrado = df_filtrado[df_filtrado['CLIENTE'].str.contains(filtro_cliente, case=False, na=False)]
    
    if filtro_container:
        df_filtrado = df_filtrado[df_filtrado['CONTAINER'].str.contains(filtro_container, case=False, na=False)]
    
    if filtro_canal != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['CANAL RFB'] == filtro_canal]
    
    # Tabela principal
    st.subheader(f"📋 Lista de Trackings ({len(df_filtrado)} registros)")
    
    if not df_filtrado.empty:
        # Aplicar cores à tabela
        styled_df = df_filtrado.style.apply(colorir_linha, axis=1)
        st.dataframe(styled_df, use_container_width=True, height=400)
        
        # Formulário para novo registro
        with st.expander("➕ Adicionar Novo Tracking"):
            with st.form("novo_tracking"):
                col1, col2 = st.columns(2)
                
                with col1:
                    novo_cliente = st.text_input("Cliente *", placeholder="Nome do cliente...")
                    novo_container = st.text_input("Container *", placeholder="Número do container...")
                    carregamento = st.text_input("Carregamento", placeholder="DD/MM/AAAA")
                    embarque = st.text_input("Embarque Navio", placeholder="DD/MM/AAAA")
                    saida = st.text_input("Saída Navio", placeholder="DD/MM/AAAA")
                    previsao = st.text_input("Previsão Chegada Paranaguá", placeholder="DD/MM/AAAA")
                
                with col2:
                    chegada = st.text_input("Chegada Paranaguá", placeholder="DD/MM/AAAA")
                    canal_rfb = st.selectbox("Canal RFB", ['', 'VERDE', 'VERMELHO'])
                    liberacao = st.text_input("Liberação Paranaguá", placeholder="DD/MM/AAAA")
                    chegada_py = st.text_input("Chegada Ciudad del Este PY", placeholder="DD/MM/AAAA")
                    descarregamento = st.text_input("Descarregamento", placeholder="DD/MM/AAAA")
                
                submitted = st.form_submit_button("💾 Salvar Tracking", type="primary")
                
                if submitted:
                    if not novo_cliente or not novo_container:
                        st.error("❌ Cliente e Container são obrigatórios!")
                    else:
                        novo_registro = {
                            'CLIENTE': novo_cliente,
                            'CONTAINER': novo_container,
                            'CARREGAMENTO': carregamento,
                            'EMBARQUE NAVIO': embarque,
                            'SAIDA NAVIO': saida,
                            'PREVISAO CHEGADA PARANAGUA': previsao,
                            'CHEGADA PARANAGUA': chegada,
                            'CANAL RFB': canal_rfb,
                            'LIBERAÇAO PARANAGUA': liberacao,
                            'CHEGADA CIUDAD DEL ESTE PY': chegada_py,
                            'DESCARREGAMENTO': descarregamento
                        }
                        
                        novo_df = pd.DataFrame([novo_registro])
                        st.session_state.df_tracking = pd.concat([st.session_state.df_tracking, novo_df], ignore_index=True)
                        st.success("✅ Tracking adicionado com sucesso!")
                        st.rerun()
        
        # Edição de registros
        with st.expander("✏️ Editar/Excluir Tracking"):
            if not df_filtrado.empty:
                # Seletor de registro para editar
                opcoes_edicao = [f"{row['CLIENTE']} - {row['CONTAINER']}" for _, row in df_filtrado.iterrows()]
                registro_selecionado = st.selectbox("Selecione o registro para editar:", opcoes_edicao)
                
                if registro_selecionado:
                    # Encontrar o índice do registro selecionado
                    idx_selecionado = df_filtrado.index[df_filtrado.apply(lambda x: f"{x['CLIENTE']} - {x['CONTAINER']}" == registro_selecionado, axis=1)].tolist()[0]
                    registro = st.session_state.df_tracking.loc[idx_selecionado]
                    
                    col1, col2 = st.columns([3, 1])
                    
                    with col1:
                        st.write(f"**Editando:** {registro['CLIENTE']} - {registro['CONTAINER']}")
                    
                    with col2:
                        if st.button("🗑️ Excluir Registro", type="secondary"):
                            st.session_state.df_tracking = st.session_state.df_tracking.drop(idx_selecionado).reset_index(drop=True)
                            st.success("🗑️ Registro excluído!")
                            st.rerun()
                    
                    # Formulário de edição
                    with st.form("editar_tracking"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            edit_cliente = st.text_input("Cliente", value=registro['CLIENTE'])
                            edit_container = st.text_input("Container", value=registro['CONTAINER'])
                            edit_carregamento = st.text_input("Carregamento", value=registro['CARREGAMENTO'])
                            edit_embarque = st.text_input("Embarque Navio", value=registro['EMBARQUE NAVIO'])
                            edit_saida = st.text_input("Saída Navio", value=registro['SAIDA NAVIO'])
                            edit_previsao = st.text_input("Previsão Chegada Paranaguá", value=registro['PREVISAO CHEGADA PARANAGUA'])
                        
                        with col2:
                            edit_chegada = st.text_input("Chegada Paranaguá", value=registro['CHEGADA PARANAGUA'])
                            edit_canal = st.selectbox("Canal RFB", ['', 'VERDE', 'VERMELHO'], 
                                                    index=['', 'VERDE', 'VERMELHO'].index(registro['CANAL RFB']) if registro['CANAL RFB'] in ['', 'VERDE', 'VERMELHO'] else 0)
                            edit_liberacao = st.text_input("Liberação Paranaguá", value=registro['LIBERAÇAO PARANAGUA'])
                            edit_chegada_py = st.text_input("Chegada Ciudad del Este PY", value=registro['CHEGADA CIUDAD DEL ESTE PY'])
                            edit_descarregamento = st.text_input("Descarregamento", value=registro['DESCARREGAMENTO'])
                        
                        submitted_edit = st.form_submit_button("💾 Salvar Alterações", type="primary")
                        
                        if submitted_edit:
                            if not edit_cliente or not edit_container:
                                st.error("❌ Cliente e Container são obrigatórios!")
                            else:
                                # Atualizar o registro
                                st.session_state.df_tracking.loc[idx_selecionado] = [
                                    edit_cliente, edit_container, edit_carregamento, edit_embarque,
                                    edit_saida, edit_previsao, edit_chegada, edit_canal,
                                    edit_liberacao, edit_chegada_py, edit_descarregamento
                                ]
                                st.success("✅ Registro atualizado com sucesso!")
                                st.rerun()
    else:
        st.info("🔍 Nenhum registro encontrado com os filtros aplicados.")
    
    # Rodapé com informações
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("### 🟢 Canal Verde")
        st.write("✅ Liberação automática")
        st.write("⚡ Processo rápido")
        st.write("🤖 Sem conferência física")
    
    with col2:
        st.markdown("### 🔴 Canal Vermelho")
        st.write("🔍 Conferência física")
        st.write("⏳ Processo mais lento")
        st.write("👮 Fiscalização rigorosa")
    
    with col3:
        st.markdown("### 🌐 Sites Úteis")
        st.markdown("🔗 [Portal Siscomex](https://portalunico.siscomex.gov.br)")
        st.markdown("🔗 [Receita Federal](https://www.gov.br/receitafederal)")
        st.markdown("🔗 [Porto Paranaguá](https://www.portoparanagua.com.br)")
    
    # Alertas e notificações
    st.markdown("---")
    
    # Verificar containers com Canal Vermelho há muito tempo
    if not st.session_state.df_tracking.empty:
        containers_vermelho = st.session_state.df_tracking[
            st.session_state.df_tracking['CANAL RFB'] == 'VERMELHO'
        ]
        
        if not containers_vermelho.empty:
            st.warning(f"⚠️ **Atenção:** {len(containers_vermelho)} container(s) no Canal Vermelho precisam de acompanhamento!")
            
            with st.expander("Ver Containers no Canal Vermelho"):
                for _, row in containers_vermelho.iterrows():
                    st.write(f"🔴 **{row['CLIENTE']}** - Container: {row['CONTAINER']}")
    
    # Containers sem canal definido
    containers_pendentes = st.session_state.df_tracking[
        st.session_state.df_tracking['CANAL RFB'].isin(['', None])
    ]
    
    if not containers_pendentes.empty:
        st.info(f"📋 **Info:** {len(containers_pendentes)} container(s) aguardando definição de canal RFB.")

# Página de configurações
def pagina_configuracoes():
    st.header("⚙️ Configurações do Sistema")
    
    # Dados da empresa
    st.subheader("🏢 Dados da Empresa")
    with st.form("config_empresa"):
        nome_empresa = st.text_input("Nome da Empresa", value=DADOS_EMPRESA['nome'])
        endereco = st.text_input("Endereço", value=DADOS_EMPRESA['endereco'])
        cidade = st.text_input("Cidade", value=DADOS_EMPRESA['cidade'])
        telefone = st.text_input("Telefone", value=DADOS_EMPRESA['telefone'])
        email = st.text_input("Email", value=DADOS_EMPRESA['email'])
        cnpj = st.text_input("CNPJ", value=DADOS_EMPRESA['cnpj'])
        
        if st.form_submit_button("💾 Salvar Configurações"):
            # Aqui você pode implementar a lógica para salvar as configurações
            st.success("✅ Configurações salvas com sucesso!")
    
    # Backup e restauração
    st.subheader("💾 Backup e Restauração")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Fazer Backup:**")
        if st.button("📤 Gerar Backup Completo", type="primary"):
            if not st.session_state.df_tracking.empty:
                buffer = exportar_excel(st.session_state.df_tracking)
                st.download_button(
                    label="💾 Baixar Backup",
                    data=buffer,
                    file_name=f"backup_brix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("⚠️ Nenhum dado para fazer backup.")
    
    with col2:
        st.write("**Limpar Dados:**")
        if st.button("🗑️ Limpar Todos os Dados", type="secondary"):
            if st.checkbox("Confirmo que quero limpar todos os dados"):
                st.session_state.df_tracking = pd.DataFrame(columns=COLUNAS)
                st.success("🗑️ Todos os dados foram removidos!")
                st.rerun()

# Página de relatórios
def pagina_relatorios():
    st.header("📊 Relatórios e Estatísticas")
    
    if st.session_state.df_tracking.empty:
        st.warning("⚠️ Nenhum dado disponível para relatórios.")
        return
    
    # Estatísticas gerais
    st.subheader("📈 Estatísticas Gerais")
    
    col1, col2, col3, col4 = st.columns(4)
    
    total = len(st.session_state.df_tracking)
    verde = len(st.session_state.df_tracking[st.session_state.df_tracking['CANAL RFB'] == 'VERDE'])
    vermelho = len(st.session_state.df_tracking[st.session_state.df_tracking['CANAL RFB'] == 'VERMELHO'])
    pendentes = total - verde - vermelho
    
    with col1:
        st.metric("📦 Total", total)
    with col2:
        st.metric("🟢 Verde", verde, f"{(verde/total*100):.1f}%")
    with col3:
        st.metric("🔴 Vermelho", vermelho, f"{(vermelho/total*100):.1f}%")
    with col4:
        st.metric("⏳ Pendentes", pendentes, f"{(pendentes/total*100):.1f}%")
    
    # Gráficos detalhados
    col1, col2 = st.columns(2)
    
    with col1:
        # Evolução temporal (se houver datas)
        st.subheader("📅 Timeline de Processos")
        
        # Análise de clientes
        cliente_stats = st.session_state.df_tracking.groupby('CLIENTE').agg({
            'CONTAINER': 'count',
            'CANAL RFB': lambda x: (x == 'VERDE').sum()
        }).rename(columns={'CONTAINER': 'Total', 'CANAL RFB': 'Verde'})
        
        cliente_stats['Vermelho'] = cliente_stats['Total'] - cliente_stats['Verde']
        cliente_stats = cliente_stats.sort_values('Total', ascending=True).tail(10)
        
        fig_clientes = go.Figure()
        fig_clientes.add_trace(go.Bar(
            name='Verde', 
            y=cliente_stats.index, 
            x=cliente_stats['Verde'],
            orientation='h',
            marker_color='#27ae60'
        ))
        fig_clientes.add_trace(go.Bar(
            name='Vermelho', 
            y=cliente_stats.index, 
            x=cliente_stats['Vermelho'],
            orientation='h',
            marker_color='#e74c3c'
        ))
        
        fig_clientes.update_layout(
            title="📊 Containers por Cliente e Canal",
            barmode='stack',
            height=400
        )
        st.plotly_chart(fig_clientes, use_container_width=True)
    
    with col2:
        # Status de completude
        st.subheader("✅ Status de Completude")
        
        # Calcular campos preenchidos por registro
        campos_importantes = ['CARREGAMENTO', 'EMBARQUE NAVIO', 'CHEGADA PARANAGUA', 'LIBERAÇAO PARANAGUA', 'DESCARREGAMENTO']
        
        completude = []
        for _, row in st.session_state.df_tracking.iterrows():
            preenchidos = sum(1 for campo in campos_importantes if row[campo] and str(row[campo]).strip())
            percentual = (preenchidos / len(campos_importantes)) * 100
            completude.append(percentual)
        
        # Histograma de completude
        fig_completude = px.histogram(
            x=completude,
            bins=20,
            title="📋 Distribuição de Completude dos Registros",
            labels={'x': 'Percentual de Completude (%)', 'y': 'Quantidade de Registros'},
            color_discrete_sequence=['#3498db']
        )
        fig_completude.update_layout(height=400)
        st.plotly_chart(fig_completude, use_container_width=True)
    
    # Tabela resumo por cliente
    st.subheader("📋 Resumo por Cliente")
    
    resumo_cliente = st.session_state.df_tracking.groupby('CLIENTE').agg({
        'CONTAINER': 'count',
        'CANAL RFB': [
            lambda x: (x == 'VERDE').sum(),
            lambda x: (x == 'VERMELHO').sum(),
            lambda x: (x.isin(['', None]) | x.isna()).sum()
        ]
    }).round(2)
    
    resumo_cliente.columns = ['Total', 'Verde', 'Vermelho', 'Pendente']
    resumo_cliente = resumo_cliente.reset_index()
    
    st.dataframe(resumo_cliente, use_container_width=True)

# Menu principal
def menu_principal():
    # Menu lateral
    with st.sidebar:
        st.title("🚢 BRIX Tracking")
        
        opcao = st.radio(
            "Escolha uma opção:",
            ["📊 Dashboard", "📋 Relatórios", "⚙️ Configurações", "❓ Ajuda"]
        )
        
        st.markdown("---")
        st.markdown("### 📱 Acesso Móvel")
        st.markdown("Este sistema funciona perfeitamente no celular!")
        
        st.markdown("### 🔗 Links Úteis")
        st.markdown("🌐 [Portal Siscomex](https://portalunico.siscomex.gov.br)")
        st.markdown("🌐 [Receita Federal](https://www.gov.br/receitafederal)")
        
        st.markdown("---")
        st.markdown(f"**📅 Última atualização:**  \n{datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Conteúdo principal baseado na seleção
    if opcao == "📊 Dashboard":
        main()
    elif opcao == "📋 Relatórios":
        pagina_relatorios()
    elif opcao == "⚙️ Configurações":
        pagina_configuracoes()
    elif opcao == "❓ Ajuda":
        pagina_ajuda()

# Página de ajuda
def pagina_ajuda():
    st.header("❓ Ajuda e Documentação")
    
    with st.expander("🚀 Como Começar"):
        st.markdown("""
        ### 1. Carregar Dados
        - Use "Carregar Dados de Exemplo" na barra lateral
        - Ou importe seu arquivo Excel existente
        
        ### 2. Adicionar Novos Trackings
        - Clique em "Adicionar Novo Tracking" no dashboard
        - Preencha pelo menos Cliente e Container (obrigatórios)
        
        ### 3. Filtrar e Buscar
        - Use os filtros na parte superior do dashboard
        - Filtros funcionam em tempo real
        """)
    
    with st.expander("🎨 Sistema de Cores"):
        st.markdown("""
        ### 🟢 Verde
        - Canal RFB: Liberação automática
        - Processo mais rápido
        - Sem conferência física necessária
        
        ### 🔴 Vermelho
        - Canal RFB: Conferência física obrigatória
        - Processo mais demorado
        - Fiscalização rigorosa
        
        ### ⚪ Sem cor
        - Canal RFB ainda não definido
        - Aguardando classificação
        """)
    
    with st.expander("📱 Acesso Mobile"):
        st.markdown("""
        ### 📱 Funcionamento no Celular
        - Interface responsiva
        - Funciona em qualquer dispositivo
        - Mesmas funcionalidades da versão desktop
        
        ### 💡 Dicas para Mobile
        - Use filtros para encontrar rapidamente
        - Tabela pode ser rolada horizontalmente
        - Formulários se adaptam ao tamanho da tela
        """)
    
    with st.expander("🔧 Solução de Problemas"):
        st.markdown("""
        ### ❌ Problemas Comuns
        
        **Arquivo não carrega:**
        - Verifique se é um arquivo Excel (.xlsx ou .xls)
        - Confirme se as colunas estão corretas
        
        **Dados não aparecem:**
        - Verifique os filtros aplicados
        - Clique em "Limpar Filtros"
        
        **Sistema lento:**
        - Normal com muitos dados
        - Use filtros para reduzir a quantidade exibida
        """)
    
    with st.expander("🌐 Sites Governamentais"):
        st.markdown("""
        ### 📋 Principais Sites
        
        **Portal Único Siscomex:**
        - URL: https://portalunico.siscomex.gov.br
        - Declarações de importação/exportação
        - ⚠️ Pode apresentar lentidão
        
        **Receita Federal:**
        - URL: https://www.gov.br/receitafederal
        - Consulta de liberações
        - Status de processos
        
        **Porto de Paranaguá:**
        - URL: https://www.portoparanagua.com.br
        - Chegada de navios
        - Status de cargas
        
        ### 💡 Dicas
        - Sites podem ficar lentos às segundas-feiras
        - Evite acessar no final do mês
        - Tenha paciência - pode demorar para carregar
        """)

if __name__ == "__main__":
    menu_principal()